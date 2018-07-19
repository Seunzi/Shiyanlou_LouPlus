# -*- coding: utf-8 -*-

from openpyxl import load_workbook
from openpyxl import Workbook
from openpyxl.utils.dataframe import dataframe_to_rows
import datetime
from itertools import islice
from pandas import DataFrame

def ws2df(ws):
    data = ws.values
    cols = next(data)
    data = list(data)
    data = [islice(r,0,None) for r in data]
    df = DataFrame(data,columns=cols)
    return df

def combine():
    wb = load_workbook(filename = 'courses.xlsx')
    ws_stu = wb['students']
    ws_time = wb['time']
    ws_combine = wb.create_sheet(title="combine")
    df_stu = ws2df(ws_stu)
    df_time = ws2df(ws_time)
    df_combine = df_stu.merge(df_time)

    for r in dataframe_to_rows(df_combine, index=False, header=True):
        ws_combine.append(r)
    
    wb.save(filename = 'courses.xlsx')

def split():
    wb = load_workbook(filename = 'courses.xlsx')
    ws_combine = wb['combine']
    df_combine = ws2df(ws_combine)
    min_year = df_combine['创建时间'].min().year
    max_year = df_combine['创建时间'].max().year
    for y in range(min_year,(max_year + 1)):
        df_year = df_combine[(df_combine['创建时间'] >= str(y)) & (df_combine['创建时间'] < str(y+1))]
        wb = Workbook()
        dest_filename = "{}.xlsx".format(y)
        ws = wb.active
        ws.title = "{}".format(y)
        for r in dataframe_to_rows(df_year, index=False, header=True):
            ws.append(r)
        wb.save(filename = dest_filename)

if __name__ == '__main__':
    combine()
    split()

